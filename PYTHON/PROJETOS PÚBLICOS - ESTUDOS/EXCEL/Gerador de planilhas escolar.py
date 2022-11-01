import openpyxl
import pyfiglet

i = 0

def cadastro():
    bem_vindo_text = pyfiglet.figlet_format("BEM VINDO", font = "slant"  )
    print(bem_vindo_text)  
    print(70*'*')
    print("\n")
    print("============ BEM VINDO AO GERADOR DE PLANILHAS ESCOLAR =============")
    print("\n")
    print(70*'*')
    planilha = openpyxl.Workbook()
    pagina = input("\nDIGITE O NOME DA PÁGINA: ")
    print("\n")
    print(70*'*')
    planilha.create_sheet(pagina)
    pagina_cad = planilha[pagina]
    pagina_cad.append(['ALUNO', 'MATRICULA', 'TELEFONE', 'MEDIA'])
    numero_rep = int(input("DIGITE QUANTOS ALUNOS DESEJAS CADASTRAR: "))
    global i
    while i != numero_rep:
        aluno = input("\nDigite o nome do aluno: ")
        matricula = int(input("Digite a matricula do aluno: "))
        telefone = int(input("Digite o número do aluno: "))
        nota_1 = int(input("Digite a nota da primeira unidade do aluno: "))
        nota_2 = int(input("Digite a nota da segunda unidade do aluno: ")) 
        nota_3 = int(input("Digite a nota da terceira unidade do aluno: "))
        nota_4 = int(input("Digite a nota da quarta unidade do aluno: "))
        media = nota_1 + nota_2 + nota_3 + nota_4
        media_final = media / 4
        pagina_cad.append([aluno, matricula, telefone, media_final])
        planilha.save('Notas alunos.xlsx')
        i += 1
    def menu(): 
        print("\n")
        print(70*'*')
        print(10*" ", "SELECIONE UMA DAS OPÇÕES ABAIXO", 10*" ")
        print(70*'*')
        print('\n1 - ALTERAR NOME')
        print('\n2 - ALTERAR MATRICULA')
        print('\n3 - ALTERAR TELEFONE')
        print('\n4 - ALTERAR MÉDIA')
        print('\n5 - LISTAR TODAS AS INFORMAÇÕES')
        print('\n6 - SAIR')
        print(70*'*')
        option = int(input("\nDigite qual opção desejas: "))

        if option == 1:
            def alterar_nome():
                alter_n = input("Digite qual aluno desejas alterar o nome: ")
                if alter_n == aluno:
                    nome_alter = input("Digite o novo nome: ")
                    for rows in pagina_cad.iter_rows(min_row=2, max_row=numero_rep+1):
                        for cell in rows:
                            if cell.value == aluno:
                                cell.value = nome_alter
                    planilha.save('Notas alunos.xlsx')
            alterar_nome()
            menu()
                
        elif option == 2:
            def alterar_matricula():
                alterar_mat = input("Digite qual matricula desejas alterar: ")
                if alterar_mat == matricula:
                    mat_alter = input("Digite a nova matricula: ")
                    for rows in pagina_cad.iter_rows(min_row=2, max_row=numero_rep+1):
                        for cell in rows:
                            if cell.value == matricula:
                                cell.value = mat_alter
                    planilha.save('Notas alunos.xlsx')
            alterar_matricula()
            menu()

        elif option == 3:
            def alterar_telefone():
                alterar_tel = input("Digite qual telefone desejas alterar: ")
                if alterar_tel == telefone:
                    tel_alter = input("Digite a nova matricula: ")
                    for rows in pagina_cad.iter_rows(min_row=2, max_row=numero_rep+1):
                        for cell in rows:
                            if cell.value == telefone:
                                cell.value = tel_alter
                    planilha.save('Notas alunos.xlsx')
            alterar_telefone()
            menu()

        elif option == 4:
            def alterar_media():
                alterar_media = input("Digite qual media desejas alterar: ")
                if alterar_media == media_final:
                    nota_1 = int(input("Digite a nota da primeira unidade do aluno: "))
                    nota_2 = int(input("Digite a nota da segunda unidade do aluno: ")) 
                    nota_3 = int(input("Digite a nota da terceira unidade do aluno: "))
                    nota_4 = int(input("Digite a nota da quarta unidade do aluno: "))
                    med_alter = nota_1 + nota_2 + nota_3 + nota_4
                    med_alter_final = med_alter / 4
                    for rows in pagina_cad.iter_rows(min_row=2, max_row=numero_rep+1):
                        for cell in rows:
                            if cell.value == media_final:
                                cell.value = med_alter_final
                    planilha.save('Notas alunos.xlsx')
            alterar_media()
            menu()

        elif option == 5:
            for rows in pagina_cad.iter_rows(min_row=2, max_row=numero_rep+1):
                for cell in rows:
                    print(cell.value)
            menu()

        elif option == 6:
            print(70*'*')
            print("Tem certeza de que deseja sair?")
            decision = input("Se sim digite (s), se não digite (n): ")
            if decision =='s':
                def sair():
                    return

    menu()

cadastro()