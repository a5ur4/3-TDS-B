import openpyxl
i = 0
def cadastro():
    planilha = openpyxl.Workbook()
    pagina = input("DIGITE O NOME DA PÁGINA: ")
    planilha.create_sheet(pagina)
    pagina_cad = planilha[pagina]
    pagina_cad.append(['ALUNO', 'MATRICULA', 'TELEFONE', 'NOTA'])
    numero_rep = int(input("DIGITE QUANTOS ALUNOS DESEJAS CADASTRAR: "))
    global i
    while i != numero_rep:
        aluno = input("Digite o nome do aluno: ")
        matricula = int(input("Digite a matricula do aluno: "))
        telefone = int(input("Digite o número do aluno: "))
        nota = int(input("Digite a nota do aluno: "))
        pagina_cad.append([aluno, matricula, telefone, nota])
        planilha.save('Notas alunos.xlsx')
        i += 1
        if i == numero_rep: break
    mostrar = input("Deseja mostrar os alunos cadastrados no terminal? se sim digite (s), caso não apenas pressione enter: ")
    if mostrar == 's':
        for rows in pagina_cad.iter_rows(min_row=2, max_row=numero_rep+1):
            for cell in rows:
                print(cell.value)
    else:
        def sair():
            return

cadastro()
