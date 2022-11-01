#para instalar essa biblioteca, digite no terminal do VS CODE ou PY CHARM o seguinte comando 
# pip install openpyxl
import openpyxl

#criar uma planilha(book)
book = openpyxl.Workbook()
#como visualizar páginas existentes
print(book.sheetnames)
#como criar uma pagina
book.create_sheet('Frutas')
#como selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Frutas', 'Quantidades', 'Preço'])
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Maça', '3', 'R$10,90'])
frutas_page.append(['Uva', '2', 'R$15,90'])
frutas_page.append(['Abacate', '8', 'R$6,90'])
#salvar planiha
book.save('Planilha de frutas.xlsx')